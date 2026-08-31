from django.contrib import admin
from django.http import HttpResponse
import openpyxl

from .models import Quiz, Question, MCQuestion, EssayQuestion, Sitting, Choice


# ======================================================
#  COMMON EXPORT TO EXCEL FUNCTION
# ======================================================
def export_to_excel(modeladmin, request, queryset):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = modeladmin.model.__name__

    # Get model fields
    fields = [field.name for field in modeladmin.model._meta.fields]

    # Header
    for col_num, field in enumerate(fields, 1):
        worksheet.cell(row=1, column=col_num).value = field

    # Data
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field)

            if callable(value):
                value = value()
            if value is None:
                value = ""

            worksheet.cell(row=row_num, column=col_num).value = str(value)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename={modeladmin.model.__name__}.xlsx'

    workbook.save(response)
    return response

export_to_excel.short_description = " Download selected as Excel"


# ======================================================
#  INLINE CHOICES
# ======================================================
class ChoiceInline(admin.TabularInline):
    model = Choice
    extra = 2


# ======================================================
#  QUIZ ADMIN
# ======================================================
@admin.register(Quiz)
class QuizAdmin(admin.ModelAdmin):
    list_display = ('title', 'course', 'category', 'pass_mark', 'draft', 'timestamp')
    search_fields = ('title', 'course__title', 'category')
    list_filter = ('course', 'category', 'draft')
    ordering = ('-timestamp',)
    actions = [export_to_excel]


# ======================================================
#  QUESTION ADMIN
# ======================================================
@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ('short_content', 'quiz_list', 'get_type')
    search_fields = ('content', 'quiz__title')
    list_filter = ('quiz',)
    actions = [export_to_excel]

    def short_content(self, obj):
        return obj.content[:60] + "..." if len(obj.content) > 60 else obj.content
    short_content.short_description = "Question"

    def quiz_list(self, obj):
        return ", ".join([q.title for q in obj.quiz.all()])
    quiz_list.short_description = "Quizzes"

    def get_type(self, obj):
        return obj.__class__.__name__
    get_type.short_description = "Type"


# ======================================================
#  MC QUESTION ADMIN
# ======================================================
@admin.register(MCQuestion)
class MCQuestionAdmin(admin.ModelAdmin):
    list_display = ('short_content', 'quiz_list', 'choice_order')
    search_fields = ('content', 'quiz__title')
    list_filter = ('quiz', 'choice_order')
    inlines = [ChoiceInline]
    actions = [export_to_excel]

    def short_content(self, obj):
        return obj.content[:60] + "..." if len(obj.content) > 60 else obj.content
    short_content.short_description = "Question"

    def quiz_list(self, obj):
        return ", ".join([q.title for q in obj.quiz.all()])
    quiz_list.short_description = "Quizzes"


# ======================================================
#  ESSAY QUESTION ADMIN
# ======================================================
@admin.register(EssayQuestion)
class EssayQuestionAdmin(admin.ModelAdmin):
    list_display = ('short_content', 'quiz_list')
    search_fields = ('content', 'quiz__title')
    list_filter = ('quiz',)
    actions = [export_to_excel]

    def short_content(self, obj):
        return obj.content[:60] + "..." if len(obj.content) > 60 else obj.content
    short_content.short_description = "Question"

    def quiz_list(self, obj):
        return ", ".join([q.title for q in obj.quiz.all()])
    quiz_list.short_description = "Quizzes"


# ======================================================
#  CHOICE ADMIN
# ======================================================
@admin.register(Choice)
class ChoiceAdmin(admin.ModelAdmin):
    list_display = ('choice_text', 'question', 'correct')
    list_filter = ('correct',)
    search_fields = ('choice_text',)
    actions = [export_to_excel]


# ======================================================
#  SITTING ADMIN
# ======================================================
@admin.register(Sitting)
class SittingAdmin(admin.ModelAdmin):
    list_display = (
        'employee_display',
        'quiz',
        'score_display',
        'get_percent',
        'complete',
        'start',
        'end'
    )
    search_fields = ('employee__username', 'quiz__title')
    list_filter = ('quiz', 'complete')
    ordering = ('-end',)
    actions = [export_to_excel]

    readonly_fields = (
        'employee_display',
        'quiz',
        'score_display',
        'get_percent',
        'user_answers',
        'incorrect_questions',
        'complete',
        'start',
        'end'
    )

    def employee_display(self, obj):
        return f"{obj.employee.username} ({obj.employee.first_name} {obj.employee.last_name})"
    employee_display.short_description = "Employee"

    def score_display(self, obj):
        return obj.get_current_score()
    score_display.short_description = "Score"

    def get_percent(self, obj):
        return f"{obj.get_percent_correct()}%"
    get_percent.short_description = "Percentage"

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('employee', 'quiz')